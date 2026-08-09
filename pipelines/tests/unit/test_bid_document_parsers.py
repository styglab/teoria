import io
import subprocess
import zipfile

from openpyxl import Workbook

import pytest

from teoria_pipelines.document_parsers import (
    UnsupportedDocumentError,
    parse_document,
    sanitize_document_content,
)


def test_text_parser_replaces_nul_with_space() -> None:
    parser, result = parse_document(b"\xec\x9e\x85\xec\xb0\xb0\x00\xec\x9e\x90\xea\xb2\xa9", "notice.txt", "text/plain")

    assert parser == "text"
    assert result["blocks"][0]["text"] == "입찰 자격"


def test_sanitizes_legacy_parsed_document_blocks() -> None:
    content = {"blocks": [{"block_id": "p1", "text": "입찰\x00참가\x00자격"}]}

    assert sanitize_document_content(content)["blocks"][0]["text"] == "입찰 참가 자격"


def test_parses_hwpx_paragraphs() -> None:
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as archive:
        archive.writestr(
            "Contents/section0.xml",
            '<hs:sec xmlns:hs="urn:test"><hp:p xmlns:hp="urn:p"><hp:t>입찰참가자격</hp:t></hp:p></hs:sec>',
        )
    parser, result = parse_document(stream.getvalue(), "공고.hwpx", None)
    assert parser == "hwpx"
    assert result["blocks"][0]["text"] == "입찰참가자격"


def test_parses_xlsx_rows() -> None:
    workbook = Workbook()
    workbook.active.append(["조건", "중소기업"])
    stream = io.BytesIO()
    workbook.save(stream)
    parser, result = parse_document(stream.getvalue(), "조건.xlsx", None)
    assert parser == "xlsx"
    assert result["blocks"][0]["text"] == "조건 | 중소기업"


def test_xls_ole_container_uses_spreadsheet_converter(monkeypatch) -> None:
    def fake_run(command: list[str]) -> subprocess.CompletedProcess[str]:
        output_dir = command[command.index("--outdir") + 1]
        workbook = Workbook()
        workbook.active.append(["지역", "서울특별시"])
        workbook.save(f"{output_dir}/document.xlsx")
        return subprocess.CompletedProcess(command, 0, "", "")

    monkeypatch.setattr("teoria_pipelines.document_parsers._run_converter", fake_run)
    parser, result = parse_document(
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"xls", "조건.XLS", "application/vnd.ms-excel"
    )
    assert parser == "xls-libreoffice"
    assert result["blocks"][0]["text"] == "지역 | 서울특별시"


def test_zip_parses_supported_members_and_records_unsupported_members() -> None:
    workbook = Workbook()
    workbook.active.append(["자격", "중소기업"])
    sheet = io.BytesIO()
    workbook.save(sheet)
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as archive:
        archive.writestr("조건.xlsx", sheet.getvalue())
        archive.writestr("참고.bin", b"unknown")
    parser, result = parse_document(stream.getvalue(), "첨부.zip", "application/zip")
    assert parser == "zip"
    assert result["blocks"][0]["section"] == "조건.xlsx"
    assert result["requires_review"] is True
    assert result["unavailable_members"][0]["file_name"] == "참고.bin"


def test_pdf_without_text_is_deferred() -> None:
    from pypdf import PdfWriter

    stream = io.BytesIO()
    writer = PdfWriter()
    writer.add_blank_page(width=100, height=100)
    writer.write(stream)
    with pytest.raises(UnsupportedDocumentError, match="pdf_text_unavailable"):
        parse_document(stream.getvalue(), "도면.pdf", "application/pdf")


def test_rejects_hwp_extension_without_ole_container() -> None:
    with pytest.raises(UnsupportedDocumentError, match="invalid_hwp_container"):
        parse_document(b"not an hwp", "공고.hwp", "application/octet-stream")


def test_hwp_uses_pyhwp_before_libreoffice(monkeypatch) -> None:
    calls: list[str] = []

    def fake_run(command: list[str]) -> subprocess.CompletedProcess[str]:
        calls.append(command[0])
        output = command[command.index("--output") + 1]
        with open(output, "w", encoding="utf-8") as stream:
            stream.write("입찰참가자격\n\n중소기업 확인서")
        return subprocess.CompletedProcess(command, 0, "", "")

    monkeypatch.setattr("teoria_pipelines.document_parsers._run_converter", fake_run)
    parser, result = parse_document(
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"data", "공고.hwp", None
    )

    assert parser == "hwp-pyhwp"
    assert calls == ["hwp5txt"]
    assert len(result["blocks"]) == 2
    assert result["diagnostics"]["extracted_char_count"] > 0


def test_ole_signature_takes_precedence_over_incorrect_hwpx_extension(monkeypatch) -> None:
    def fake_run(command: list[str]) -> subprocess.CompletedProcess[str]:
        output = command[command.index("--output") + 1]
        with open(output, "w", encoding="utf-8") as stream:
            stream.write("실제 바이너리 HWP")
        return subprocess.CompletedProcess(command, 0, "", "")

    monkeypatch.setattr("teoria_pipelines.document_parsers._run_converter", fake_run)
    parser, result = parse_document(
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"data", "잘못된확장자.hwpx", None
    )

    assert parser == "hwp-pyhwp"
    assert result["blocks"][0]["text"] == "실제 바이너리 HWP"


def test_hwp_falls_back_to_libreoffice(monkeypatch) -> None:
    calls: list[str] = []

    def fake_run(command: list[str]) -> subprocess.CompletedProcess[str]:
        calls.append(command[0])
        if command[0] == "hwp5txt":
            return subprocess.CompletedProcess(command, 1, "", "invalid hwp")
        output_dir = command[command.index("--outdir") + 1]
        from pypdf import PdfWriter

        writer = PdfWriter()
        writer.add_blank_page(width=100, height=100)
        with open(f"{output_dir}/document.pdf", "wb") as stream:
            writer.write(stream)
        return subprocess.CompletedProcess(command, 0, "", "")

    monkeypatch.setattr("teoria_pipelines.document_parsers._run_converter", fake_run)
    with pytest.raises(UnsupportedDocumentError, match="hwp_requires_ocr"):
        parse_document(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"data", "공고.hwp", None)
    assert calls == ["hwp5txt", "soffice"]
