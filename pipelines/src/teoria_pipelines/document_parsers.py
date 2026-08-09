from __future__ import annotations

import io
import re
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader


PARSER_VERSION = "2.1.1"

_OLE_COMPOUND_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class UnsupportedDocumentError(ValueError):
    pass


def parse_document(value: bytes, file_name: str | None, media_type: str | None) -> tuple[str, dict[str, Any]]:
    suffix = Path(file_name or "").suffix.lower()
    if value.startswith(b"%PDF") or suffix == ".pdf":
        parsed = _parse_pdf(value)
        if not parsed["blocks"]:
            raise UnsupportedDocumentError("pdf_text_unavailable")
        return "pdf", parsed
    if _is_hwpx(value):
        return "hwpx", _parse_hwpx(value)
    if suffix == ".xls":
        return "xls-libreoffice", _parse_xls(value)
    if value.startswith(_OLE_COMPOUND_MAGIC):
        return _parse_hwp(value)
    if suffix == ".hwpx":
        raise UnsupportedDocumentError("invalid_hwpx_container")
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx", _parse_xlsx(value)
    if suffix == ".zip":
        return "zip", _parse_zip(value)
    if suffix == ".hwp":
        raise UnsupportedDocumentError("invalid_hwp_container")
    if (media_type or "").startswith("text/"):
        text = value.decode("utf-8", errors="replace")
        return "text", {"blocks": [_block("b1", None, None, "paragraph", text)]}
    raise UnsupportedDocumentError(f"unsupported_document_type:{suffix or media_type or 'unknown'}")


def _block(block_id: str, page: int | None, section: str | None,
           block_type: str, text: str) -> dict[str, Any]:
    return {"block_id": block_id, "page": page, "section": section,
            "type": block_type, "text": _sanitize_text(text)}


def sanitize_document_content(content: dict[str, Any]) -> dict[str, Any]:
    """Remove PostgreSQL-incompatible NULs from current and legacy parsed blocks."""
    for block in content.get("blocks", []):
        if isinstance(block.get("text"), str):
            block["text"] = _sanitize_text(block["text"])
    return content


def _sanitize_text(text: str) -> str:
    return text.replace("\x00", " ").strip()


def _parse_pdf(value: bytes) -> dict[str, Any]:
    blocks = []
    for index, page in enumerate(PdfReader(io.BytesIO(value)).pages, start=1):
        text = page.extract_text() or ""
        if text.strip():
            blocks.append(_block(f"p{index}", index, None, "page", text))
    return {"blocks": blocks, "requires_ocr": not blocks}


def _is_hwpx(value: bytes) -> bool:
    try:
        with zipfile.ZipFile(io.BytesIO(value)) as archive:
            return any(name.startswith("Contents/section") for name in archive.namelist())
    except zipfile.BadZipFile:
        return False


def _parse_hwpx(value: bytes) -> dict[str, Any]:
    blocks = []
    with zipfile.ZipFile(io.BytesIO(value)) as archive:
        sections = sorted(name for name in archive.namelist()
                          if name.startswith("Contents/section") and name.endswith(".xml"))
        for section_index, name in enumerate(sections, start=1):
            root = ElementTree.fromstring(archive.read(name))
            for paragraph_index, paragraph in enumerate(root.iter(), start=1):
                if not paragraph.tag.endswith("}p"):
                    continue
                text = "".join(node.text or "" for node in paragraph.iter()
                               if node.tag.endswith("}t")).strip()
                if text:
                    blocks.append(_block(f"s{section_index}p{paragraph_index}", None,
                                         f"section-{section_index}", "paragraph", text))
    return {"blocks": blocks}


def _parse_xlsx(value: bytes) -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(value), read_only=True, data_only=True)
    except (TypeError, ValueError):
        return _parse_xlsx_xml(value)
    blocks = []
    for sheet in workbook.worksheets:
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                blocks.append(_block(f"{sheet.title}!{row_index}", None, sheet.title,
                                     "table_row", " | ".join(cells)))
    return {"blocks": blocks}


def _parse_xlsx_xml(value: bytes) -> dict[str, Any]:
    """Read cell values even when malformed optional styles break openpyxl."""
    blocks: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(value)) as archive:
            shared_strings: list[str] = []
            if "xl/sharedStrings.xml" in archive.namelist():
                root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
                shared_strings = [
                    "".join(node.text or "" for node in item.iter() if node.tag.endswith("}t"))
                    for item in root if item.tag.endswith("}si")
                ]

            sheets = sorted(
                name for name in archive.namelist()
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            )
            for sheet_index, name in enumerate(sheets, start=1):
                root = ElementTree.fromstring(archive.read(name))
                for row in (node for node in root.iter() if node.tag.endswith("}row")):
                    cells: list[str] = []
                    for cell in (node for node in row if node.tag.endswith("}c")):
                        cell_type = cell.attrib.get("t")
                        if cell_type == "inlineStr":
                            text = "".join(
                                node.text or "" for node in cell.iter()
                                if node.tag.endswith("}t")
                            )
                        else:
                            value_node = next(
                                (node for node in cell if node.tag.endswith("}v")), None
                            )
                            text = value_node.text if value_node is not None else ""
                            if cell_type == "s" and text:
                                index = int(text)
                                text = shared_strings[index] if index < len(shared_strings) else text
                        if text and text.strip():
                            cells.append(text.strip())
                    if cells:
                        row_number = row.attrib.get("r", str(len(blocks) + 1))
                        blocks.append(_block(
                            f"sheet{sheet_index}!{row_number}", None,
                            f"sheet-{sheet_index}", "table_row", " | ".join(cells),
                        ))
    except (zipfile.BadZipFile, ElementTree.ParseError, KeyError, ValueError) as exc:
        raise UnsupportedDocumentError("invalid_xlsx_container") from exc
    return {"blocks": blocks}


def _parse_xls(value: bytes) -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        source = root / "document.xls"
        source.write_bytes(value)
        profile = root / "libreoffice-profile"
        process = _run_converter([
            "soffice", "--headless", f"-env:UserInstallation={profile.as_uri()}",
            "--convert-to", "xlsx", "--outdir", str(root), str(source),
        ])
        target = root / "document.xlsx"
        if process.returncode or not target.exists() or target.stat().st_size == 0:
            raise UnsupportedDocumentError("xls_conversion_failed")
        return _parse_xlsx(target.read_bytes())


def _parse_zip(value: bytes) -> dict[str, Any]:
    blocks: list[dict[str, Any]] = []
    unavailable: list[dict[str, str]] = []
    total_size = 0
    try:
        with zipfile.ZipFile(io.BytesIO(value)) as archive:
            members = [item for item in archive.infolist() if not item.is_dir()]
            if len(members) > 50:
                raise UnsupportedDocumentError("zip_member_limit_exceeded")
            for member_index, member in enumerate(members, start=1):
                total_size += member.file_size
                if member.file_size > 20 * 1024 * 1024 or total_size > 100 * 1024 * 1024:
                    raise UnsupportedDocumentError("zip_uncompressed_size_limit_exceeded")
                member_name = Path(member.filename).name
                if not member_name or member.filename.startswith(("/", "\\")) or ".." in Path(member.filename).parts:
                    unavailable.append({"file_name": member.filename, "error_code": "unsafe_archive_path"})
                    continue
                if Path(member_name).suffix.lower() == ".zip":
                    unavailable.append({"file_name": member_name, "error_code": "nested_zip_unsupported"})
                    continue
                try:
                    _, parsed = parse_document(archive.read(member), member_name, None)
                except (UnsupportedDocumentError, RuntimeError) as exc:
                    unavailable.append({"file_name": member_name, "error_code": str(exc)})
                    continue
                for block_index, block in enumerate(parsed.get("blocks", []), start=1):
                    blocks.append({
                        **block,
                        "block_id": f"z{member_index}b{block_index}",
                        "section": member_name,
                    })
    except zipfile.BadZipFile as exc:
        raise UnsupportedDocumentError("invalid_zip_container") from exc
    if not blocks:
        raise UnsupportedDocumentError("zip_text_unavailable")
    return {"blocks": blocks, "unavailable_members": unavailable,
            "requires_review": bool(unavailable)}


def _parse_hwp(value: bytes) -> tuple[str, dict[str, Any]]:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        source = root / "document.hwp"
        source.write_bytes(value)

        pyhwp_error: str | None = None
        try:
            parsed = _parse_hwp_with_pyhwp(source, root)
            parsed["diagnostics"] = _diagnostics(parsed, [
                {"parser": "pyhwp", "status": "parsed"},
            ])
            return "hwp-pyhwp", parsed
        except UnsupportedDocumentError as exc:
            pyhwp_error = str(exc)

        try:
            parsed = _parse_hwp_with_libreoffice(source, root)
            parsed["diagnostics"] = _diagnostics(parsed, [
                {"parser": "pyhwp", "status": "failed", "error_code": pyhwp_error},
                {"parser": "libreoffice", "status": "parsed"},
            ])
            return "hwp-libreoffice", parsed
        except UnsupportedDocumentError as exc:
            libreoffice_error = str(exc)

        if pyhwp_error in {"hwp_encrypted", "hwp_distribution_document"}:
            raise UnsupportedDocumentError(pyhwp_error)
        if libreoffice_error == "hwp_requires_ocr":
            raise UnsupportedDocumentError(libreoffice_error)
        raise UnsupportedDocumentError(
            f"hwp_all_parsers_failed:{pyhwp_error}:{libreoffice_error}"
        )


def _parse_hwp_with_pyhwp(source: Path, root: Path) -> dict[str, Any]:
    target = root / "document.txt"
    process = _run_converter(["hwp5txt", "--output", str(target), str(source)])
    if process.returncode:
        raise UnsupportedDocumentError(_classify_hwp_error(process.stderr, "pyhwp_failed"))
    if not target.exists():
        raise UnsupportedDocumentError("pyhwp_output_missing")

    text = target.read_text(encoding="utf-8", errors="replace")
    blocks = _text_blocks(text)
    if not blocks:
        raise UnsupportedDocumentError("pyhwp_empty_output")
    return {"blocks": blocks, "requires_ocr": False}


def _parse_hwp_with_libreoffice(source: Path, root: Path) -> dict[str, Any]:
    profile = root / "libreoffice-profile"
    process = _run_converter([
        "soffice",
        "--headless",
        f"-env:UserInstallation={profile.as_uri()}",
        "--convert-to", "pdf",
        "--outdir", str(root),
        str(source),
    ])
    if process.returncode:
        raise UnsupportedDocumentError(
            _classify_hwp_error(process.stderr, "libreoffice_conversion_failed")
        )
    target = root / "document.pdf"
    if not target.exists() or target.stat().st_size == 0:
        raise UnsupportedDocumentError("libreoffice_output_missing")
    parsed = _parse_pdf(target.read_bytes())
    if not parsed["blocks"]:
        raise UnsupportedDocumentError("hwp_requires_ocr")
    return parsed


def _run_converter(command: list[str]) -> subprocess.CompletedProcess[str]:
    try:
        return subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=120,
        )
    except FileNotFoundError as exc:
        raise UnsupportedDocumentError(f"converter_unavailable:{command[0]}") from exc
    except subprocess.TimeoutExpired as exc:
        raise UnsupportedDocumentError(f"converter_timeout:{command[0]}") from exc


def _classify_hwp_error(stderr: str | None, default: str) -> str:
    message = (stderr or "").casefold()
    if any(token in message for token in ("password", "encrypted", "encryption")):
        return "hwp_encrypted"
    if any(token in message for token in ("distribution", "배포용")):
        return "hwp_distribution_document"
    if any(token in message for token in ("invalid", "corrupt", "truncated", "not an ole")):
        return "hwp_corrupted"
    return default


def _text_blocks(text: str) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for page_index, page in enumerate(text.split("\f"), start=1):
        paragraphs = re.split(r"(?:\r?\n){2,}", page)
        for paragraph in paragraphs:
            normalized = "\n".join(line.rstrip() for line in paragraph.splitlines()).strip()
            if normalized:
                blocks.append(_block(
                    f"p{page_index}b{len(blocks) + 1}", page_index, None,
                    "paragraph", normalized,
                ))
    return blocks


def _diagnostics(parsed: dict[str, Any], attempts: list[dict[str, Any]]) -> dict[str, Any]:
    blocks = parsed.get("blocks", [])
    text = "\n".join(str(block.get("text", "")) for block in blocks)
    return {
        "attempts": attempts,
        "block_count": len(blocks),
        "extracted_char_count": len(text),
        "replacement_character_count": text.count("\ufffd"),
        "requires_ocr": bool(parsed.get("requires_ocr")),
    }
